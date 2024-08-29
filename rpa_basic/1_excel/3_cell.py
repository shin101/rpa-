from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

#A1셀에 1이라는 값 입력 
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"]=4
ws["B2"]=5
ws["B3"]=6

print(ws["A1"])
print(ws["A1"].value)

# row = 1,2,3, ...
# column = A(1), B(2), C(3), ... 
print(ws.cell(row=1, column=1)) # ws['A1'].value

c = ws.cell(column=3, row=1,value=10) # ws["C1"] = 10
print(c.value)

from random import *

for x in range(1,11): # 10개 row
    for y in range(1,11): # 10개 column
        ws.cell(row=x, column=y, value=randint(0,100))



wb.save("sample.xlsx")