from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff66ff"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet",2) # 2번째 index에 sheet생성

new_ws = wb['NewSheet']

# Sheet 복사
new_ws["A1"] = "Test" # A1 indicates the cell
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet "

wb.save("sample.xlsx")